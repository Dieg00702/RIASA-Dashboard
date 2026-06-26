import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# ==========================================
# 1. FUNCIÓN BASE DE LECTURA (SE QUEDA IGUAL)
# ==========================================
def cargar_tabla_excel(ruta_archivo, nombre_hoja, nombre_tabla):
    wb = load_workbook(ruta_archivo, read_only=False, data_only=True)
    ws = wb[nombre_hoja]
    if nombre_tabla in ws.tables:
        tabla_obj = ws.tables[nombre_tabla]
        rango_celdas = ws[tabla_obj.ref] 
        filas = [[celda.value for celda in fila] for fila in rango_celdas]
        wb.close()
        return pd.DataFrame(filas[1:], columns=filas[0])
    wb.close()
    return pd.DataFrame()


# ==========================================
# 2. TARJETAS INDIVIDUALES (INYECTORA, PELETIZADO, TRITURADO)
def  mostrar_indicadores():
    # --- INYECTORA (Tabla8) ---
    df = cargar_tabla_excel("Rep RIASA.xlsm", "Medidores Grafica", "Tabla8")
    if not df.empty:
        real_pza = pd.to_numeric(df.loc[0, "Real Pza"], errors="coerce") or 0
        plan_pza_min = pd.to_numeric(df.loc[0, "Plan Pza"], errors="coerce") or 0
        plan_mensual_max = pd.to_numeric(df.loc[0, "Plan mensual"], errors="coerce") or 0
        porcentaje = (real_pza / plan_pza_min * 100) if plan_pza_min > 0 else 0
        ancho_barra = min(max((real_pza / plan_mensual_max * 100), 0), 100) if plan_mensual_max > 0 else 0
        
        real_formateado = f"{real_pza:,.0f}"
        min_formateado = f"{plan_pza_min:,.0f}"
        max_formateado = f"{plan_mensual_max:,.0f}"
        porcentaje_formateado = f"{porcentaje:.1f}%"

        if porcentaje >= 95:
            color_tema = "#4cd964"
            fondo_porcentaje = "rgba(76, 217, 100, 0.15)"
        elif porcentaje >= 80:
            color_tema = "#ffcc00"
            fondo_porcentaje = "rgba(255, 204, 0, 0.15)"
        else:
            color_tema = "#ff3b30"
            fondo_porcentaje = "rgba(255, 59, 48, 0.15)"

        html_tarjeta_inyectora = f"""
        <div style="background-color: #1a1b23; border: 1px solid #2d2f39; border-radius: 12px; padding: 24px 28px; width: 100%; max-width: 460px; box-shadow: 0 4px 15px rgba(0,0,0,0.25); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">
            <p style="color: #808495; font-size: 26px; font-weight: bold; letter-spacing:.5px; margin: 0; text-transform: uppercase;">INYECTORA</p>
            <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 12px; margin-bottom: 8px;">
                <span style="color:#ffffff; font-size: 70px; font-weight: 500; margin: 5px 0 2px 0; line-height: 1.1;">{real_formateado}</span>
                <span style="color: {color_tema}; background-color: {fondo_porcentaje}; border: 1px solid {color_tema}50; padding: 6px 14px; border-radius: 20px; font-size: 15px; font-weight: bold;">{porcentaje_formateado}</span>
            </div>
            <p style="color: #636674; font-size: 14px; margin: 0; line-height: 1.4;">
                Obj. min: <span style="color: #808495;">{min_formateado}</span> &middot; Obj. max: <span style="color: #808495;">{max_formateado}</span>
            </p>
            <div style="background-color: #262730; width: 100%; border-radius: 6px; height: 8px; margin-top: 18px; margin-bottom: 4px;">
                <div style="background-color: {color_tema}; width: {ancho_barra}%; border-radius: 6px; height: 100%;"></div>
            </div>
        </div>
        """
        st.markdown(html_tarjeta_inyectora, unsafe_allow_html=True)
        st.markdown('<div style="margin-top: 40px;"></div>', unsafe_allow_html=True)

    # --- PELETIZADO (Tabla12) ---
    df = cargar_tabla_excel("Rep RIASA.xlsm", "Medidores Grafica", "Tabla12")
    if not df.empty:
        real_pza = pd.to_numeric(df.loc[0, "Real Pza"], errors="coerce") or 0
        plan_pza_min = pd.to_numeric(df.loc[0, "Plan Pza"], errors="coerce") or 0
        plan_mensual_max = pd.to_numeric(df.loc[0, "Plan mensual"], errors="coerce") or 0
        porcentaje = (real_pza / plan_pza_min * 100) if plan_pza_min > 0 else 0
        ancho_barra = min(max((real_pza / plan_mensual_max * 100), 0), 100) if plan_mensual_max > 0 else 0
        
        real_formateado = f"{real_pza:,.0f}"
        min_formateado = f"{plan_pza_min:,.0f}"
        max_formateado = f"{plan_mensual_max:,.0f}"
        porcentaje_formateado = f"{porcentaje:.1f}%"

        if porcentaje >= 95:
            color_tema = "#4cd964"
            fondo_porcentaje = "rgba(76, 217, 100, 0.15)"
        elif porcentaje >= 80:
            color_tema = "#ffcc00"
            fondo_porcentaje = "rgba(255, 204, 0, 0.15)"
        else:
            color_tema = "#ff3b30"
            fondo_porcentaje = "rgba(255, 59, 48, 0.15)"

        html_tarjeta_peletizado = f"""
        <div style="background-color: #1a1b23; border: 1px solid #2d2f39; border-radius: 12px; padding: 24px 28px; width: 100%; max-width: 460px; box-shadow: 0 4px 15px rgba(0,0,0,0.25); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">
            <p style="color: #808495; font-size: 26px; font-weight: bold; letter-spacing:.5px; margin: 0; text-transform: uppercase;">PELETIZADO</p>
            <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 12px; margin-bottom: 8px;">
                <span style="color:#ffffff; font-size: 70px; font-weight: 500; margin: 5px 0 2px 0; line-height: 1.1;">{real_formateado}</span>
                <span style="color: {color_tema}; background-color: {fondo_porcentaje}; border: 1px solid {color_tema}50; padding: 6px 14px; border-radius: 20px; font-size: 15px; font-weight: bold;">{porcentaje_formateado}</span>
            </div>
            <p style="color: #636674; font-size: 14px; margin: 0; line-height: 1.4;">
                Obj. min: <span style="color: #808495;">{min_formateado}</span> &middot; Obj. max: <span style="color: #808495;">{max_formateado}</span>
            </p>
            <div style="background-color: #262730; width: 100%; border-radius: 6px; height: 8px; margin-top: 18px; margin-bottom: 4px;">
                <div style="background-color: {color_tema}; width: {ancho_barra}%; border-radius: 6px; height: 100%;"></div>
            </div>
        </div>
        """
        st.markdown(html_tarjeta_peletizado, unsafe_allow_html=True)
        st.markdown('<div style="margin-top: 40px;"></div>', unsafe_allow_html=True)

    # --- TRITURADO (Tabla9) ---
    df = cargar_tabla_excel("Rep RIASA.xlsm", "Medidores Grafica", "Tabla9")
    if not df.empty:
        real_pza = pd.to_numeric(df.loc[0, "Real Pza"], errors="coerce") or 0
        plan_pza_min = pd.to_numeric(df.loc[0, "Plan Pza"], errors="coerce") or 0
        plan_mensual_max = pd.to_numeric(df.loc[0, "Plan mensual"], errors="coerce") or 0
        porcentaje = (real_pza / plan_pza_min * 100) if plan_pza_min > 0 else 0
        ancho_barra = min(max((real_pza / plan_mensual_max * 100), 0), 100) if plan_mensual_max > 0 else 0
        
        real_formateado = f"{real_pza:,.0f}"
        min_formateado = f"{plan_pza_min:,.0f}"
        max_formateado = f"{plan_mensual_max:,.0f}"
        porcentaje_formateado = f"{porcentaje:.1f}%"

        if porcentaje >= 95:
            color_tema = "#4cd964"
            fondo_porcentaje = "rgba(76, 217, 100, 0.15)"
        elif porcentaje >= 80:
            color_tema = "#ffcc00"
            fondo_porcentaje = "rgba(255, 204, 0, 0.15)"
        else:
            color_tema = "#ff3b30"
            fondo_porcentaje = "rgba(255, 59, 48, 0.15)"

        html_tarjeta_triturado = f"""
        <div style="background-color: #1a1b23; border: 1px solid #2d2f39; border-radius: 12px; padding: 24px 28px; width: 100%; max-width: 460px; box-shadow: 0 4px 15px rgba(0,0,0,0.25); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">
            <p style="color: #808495; font-size: 26px; font-weight: bold; letter-spacing:.5px; margin: 0; text-transform: uppercase;">TRITURADO</p>
            <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 12px; margin-bottom: 8px;">
                <span style="color:#ffffff; font-size: 70px; font-weight: 500; margin: 5px 0 2px 0; line-height: 1.1;">{real_formateado}</span>
                <span style="color: {color_tema}; background-color: {fondo_porcentaje}; border: 1px solid {color_tema}50; padding: 6px 14px; border-radius: 20px; font-size: 15px; font-weight: bold;">{porcentaje_formateado}</span>
            </div>
            <p style="color: #636674; font-size: 14px; margin: 0; line-height: 1.4;">
                Obj. min: <span style="color: #808495;">{min_formateado}</span> &middot; Obj. max: <span style="color: #808495;">{max_formateado}</span>
            </p>
            <div style="background-color: #262730; width: 100%; border-radius: 6px; height: 8px; margin-top: 18px; margin-bottom: 4px;">
                <div style="background-color: {color_tema}; width: {ancho_barra}%; border-radius: 6px; height: 100%;"></div>
            </div>
        </div>
        """
        st.markdown(html_tarjeta_triturado, unsafe_allow_html=True)
        st.markdown('<div style="margin-top: 40px;"></div>', unsafe_allow_html=True)
    def mostrar_tarjeta_vega_tdm():

        vega_real, vega_min, vega_max, vega_pct, vega_barra = 0, 0, 0, 0, 0
        tdm_real, tdm_min, tdm_max, tdm_pct, tdm_barra = 0, 0, 0, 0, 0
        
        color_vega = "#FF6B6B"
        estado_vega = "REVISAR"
        color_tdm = "#FF6B6B"
        
        # 2. INTENTAR CARGAR LAS TABLAS
    df_vega = cargar_tabla_excel("Rep RIASA.xlsm", "Medidores Grafica", "TablaVEGA") 
    df_tdm = cargar_tabla_excel("Rep RIASA.xlsm", "Medidores Grafica", "TablaTDM")
        
        # Comprobar si se cargaron correctamente y no están vacías
    if not df_vega.empty and "Real Pza" in df_vega.columns:
        vega_real = pd.to_numeric(df_vega.loc[0, "Real Pza"], errors="coerce") or 0
        vega_min = pd.to_numeric(df_vega.loc[0, "Plan Pza"], errors="coerce") or 0
        vega_max = pd.to_numeric(df_vega.loc[0, "Plan mensual"], errors="coerce") or 0
            
        vega_pct = (vega_real / vega_min * 100) if vega_min > 0 else 0
        vega_barra = min(max((vega_real / vega_max * 100), 0), 100) if vega_max > 0 else 0
            
        if vega_pct >= 95:
            color_vega = "#4ECA82"
            estado_vega = "OK"
        elif vega_pct >= 80:
                color_vega = "#ffcc00"
        else :
            color_vega = "#ff3b30"
            estado_vega = "REVISAR"

    if not df_tdm.empty and "Real Pza" in df_tdm.columns:
        tdm_real = pd.to_numeric(df_tdm.loc[0, "Real Pza"], errors="coerce") or 0
        tdm_min = pd.to_numeric(df_tdm.loc[0, "Plan Pza"], errors="coerce") or 0
        tdm_max = pd.to_numeric(df_tdm.loc[0, "Plan mensual"], errors="coerce") or 0
            
        tdm_pct = (tdm_real / tdm_min * 100) if tdm_min > 0 else 0
        tdm_barra = min(max((tdm_real / tdm_max * 100), 0), 100) if tdm_max > 0 else 0
            
        if tdm_pct >= 95:
            color_tdm = "#4ECA82"
        elif tdm_pct >= 80:
            color_tdm = "#ffcc00"
        else :
            color_tdm = "#ff3b30"

        # 3. DISEÑO HTML (Las variables ahora existen pase lo que pase)
        html_combinado = f"""
        <div style="background-color: #161922; border: 1px solid #232733; border-radius: 16px; padding: 24px; color: #FFFFFF; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 460px; box-shadow: 0 4px 15px rgba(0,0,0,0.25); margin-bottom: 20px;">
            <div style="color: #7A8599; font-size: 24px; font-weight: bold; letter-spacing: .5px; text-transform: uppercase; margin-bottom: 20px; font-weight: 600;">
                FACTURACIÓN VEGA & TDM
        </div>
            
        <div style="display: flex; justify-content: space-between; gap: 32px;">    
                <div style="flex: 1;">
                    <div style="color: #7A8599; font-size: 20px; margin-bottom: 6px;">VEGA</div>
                    <div style="color: {color_vega}; font-size: 32px; font-weight: bold; margin-bottom: 8px;">{vega_real:,.0f}</div>
                    <div style="background-color: #232733; border-radius: 10px; height: 8px; width: 100%; margin-bottom: 8px;">
                        <div style="background-color: {color_vega}; height: 100%; border-radius: 10px; width: {vega_barra}%;"></div>
                    </div>
                    <div style="color: {color_vega}; font-size: 12px; font-weight: 500;">
                        {vega_pct:.1f}% — {estado_vega}
                    </div>
        </div>
                
        <div style="flex: 1;">
                    <div style="color: #7A8599; font-size: 20px; margin-bottom: 6px;">TDM</div>
                    <div style="color: {color_tdm}; font-size: 32px; font-weight: bold; margin-bottom: 8px;">{tdm_real:,.0f}</div>
                    <div style="background-color: #232733; border-radius: 10px; height: 8px; width: 100%; margin-bottom: 8px;">
                        <div style="background-color: {color_tdm}; height: 100%; border-radius: 10px; width: {tdm_barra}%;"></div>
                    </div>
                    <div style="color: {color_tdm}; font-size: 12px; font-weight: 500;">
                        {tdm_pct:.1f}% vs {tdm_min:,.0f}
                    </div>
                </div>

        </div>
        </div>
        """
        st.markdown(html_combinado, unsafe_allow_html=True)

            