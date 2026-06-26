import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook
def  mostrar_graficas():
# Configuración de la página para emular un dashboard compacto
    st.set_page_config(layout="centered")

    # --- 1. FUNCIÓN PARA LEER EXCLUSIVAMENTE LA TABLA4 DEL EXCEL ---
    def cargar_tabla_excel(ruta_archivo, nombre_hoja, nombre_tabla):
        wb = load_workbook(ruta_archivo, read_only=False, data_only=True)
        ws = wb[nombre_hoja]
        
        if nombre_tabla in ws.tables:
            tabla_obj = ws.tables[nombre_tabla]
            rango_celdas = ws[tabla_obj.ref] 
            
            filas = [[celda.value for celda in fila] for fila in rango_celdas]
            wb.close()
            
            # El primer renglón se usa como los encabezados de columnas
            df_limpio = pd.DataFrame(filas[1:], columns=filas[0])
            return df_limpio
        else:
            wb.close()
            st.error(f"No se encontró la tabla '{nombre_tabla}'")
            return pd.DataFrame()

    # Cargar tus datos reales desde tu archivo
    df = cargar_tabla_excel("Rep RIASA.xlsm", "Medidores Grafica", "Tabla4")

    # Quitar filas basuras que tengan el mercado vacío
    df = df.dropna(subset=["MERCADO"])

    # --- 2. PROCESAMIENTO DE DATOS CON TUS ENCABEZADOS REALES ---
    # Convertir las columnas de tu Excel a números de forma segura
    df["Objetivos"] = pd.to_numeric(df["Objetivos"], errors="coerce").fillna(0)
    df["Producción"] = pd.to_numeric(df["Producción"], errors="coerce").fillna(0)

    # Calcular el porcentaje individual de cumplimiento para cada fila
    df["Porcentaje_Ind"] = (df["Producción"] / df["Objetivos"] * 100).fillna(0).round(1)

    # Totales globales para el Arco (Gauge)
    total_obj = df["Objetivos"].sum()
    total_real = df["Producción"].sum()
    total_dif = total_real - total_obj
    porcentaje_total = round((total_real / total_obj) * 100, 1) if total_obj > 0 else 0

    total_obj = df["Objetivos"].sum()
    total_real = df["Producción"].sum()
    total_dif = total_real - total_obj

    # Formateo de números a texto con comas
    real_formateado = f"{total_real:,.0f}"
    obj_formateado = f"{total_obj:,.0f}"
    dif_formateada = f"{total_dif:,.0f}"

    # Lógica del color y símbolos del Delta
    if total_dif >= 0:
        color_delta = "#4cd964"  # Verde
        simbolo_delta = "▲"
        prefijo_dif = "+"
    else:
        color_delta = "#ff3b30"  # Rojo
        simbolo_delta = "▼"
        prefijo_dif = ""         # El signo menos numérico ya viene incluido

    # ==========================================
    # 2. SEGUNDO: CREACIÓN DEL STRING HTML
    # ==========================================
    # Aquí el string lee las variables que calculamos exactamente arriba
    html_tarjeta_prod = f"""
    <div style="background-color: #1a1b23; border: 1px solid #2d2f39; border-radius: 12px; padding: 16px 20px; margin-bottom: 20px; box-shadow: 0 4px 6px rgba(0,0,0,0.15); width: 100%;">
        <p style="color: #808495; font-size: 26px; font-weight: bold; letter-spacing: 0.5px; margin: 0; text-transform: uppercase;">Producción Total</p>
        <p style="color: #ffffff; font-size: 70px; font-weight: 800; margin: 5px 0 2px 0; line-height: 1.1;">{real_formateado}</p>
        <p style="color: {color_delta}; font-size: 24px; font-weight: 600; margin: 0;">
            {simbolo_delta} {prefijo_dif}{dif_formateada} <span style="color: #808495; font-weight: normal;">vs objetivo ({obj_formateado})</span>
        </p>
    </div>
    """

    # ==========================================
    # 3. TERCERO: RENDERIZAR EN LA PÁGINA
    # ==========================================
    # Usamos 'st.' directamente a ancho completo
    st.markdown(html_tarjeta_prod, unsafe_allow_html=True)

    tarjeta_completa = st.container()
    st.markdown("""
        <style>
        /* Buscamos el contenedor específico y le aplicamos el color de la tarjeta */
        div[data-testid="stVerticalBlock"] > div:has(.tarjeta-produccion) {
            background-color: rgba(26, 27, 35, 1.0) !important;
            border-radius: 12px !important;
            
            margin-bottom: 20px !important;
            border: .5px solid #2d2f39 !important;
        }
        </style>
    """, unsafe_allow_html=True)

    # --- 2. AGREGAR EL CONTENIDO EN EL GAUGE ---
    with tarjeta_completa:
        # Esta línea vacía sirve como anclaje para el diseño CSS anterior
        st.markdown('<div class="tarjeta-produccion"></div>', unsafe_allow_html=True)
        
        # El título ahora no lleva fondo propio, usa el del contenedor general
        st.markdown("""
            <p style='font-size: 26px; 
                    color: #808495; 
                    font-weight: bold; 
                    margin: 0 0 15px 0;'>
                PRODUCCIÓN POR MERCADO
            </p>
        """, unsafe_allow_html=True)

        # --- 3. CONFIGURACIÓN DEL GRÁFICO DE PLOTLY ---
        fig = go.Figure(go.Indicator(
            mode="gauge+number",
            value=porcentaje_total,
            number={'suffix': "%", 'font': {'size': 60, 'color': 'white'}},
            gauge={
                'axis': {'range':[0, 120], 'visible': False},
                'shape': "angular", 
                'bar': {'color': "#f47920", 'thickness': 0.90}, 
                'bgcolor': "#262730",
                'borderwidth': .5       
            }
        ))

        # Quitamos los fondos propios del gráfico para que herede la transparencia del contenedor
        fig.update_layout(
            margin=dict(l=0, r=0, t=0, b=0),
            height=180,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )

        # Renderizamos el gráfico dentro de la misma tarjeta
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        
    # --- 4. DISEÑO DE INTERFAZ: BARRAS DE PROGRESO DINÁMICAS ---
    with tarjeta_completa:
        for index, row in df.iterrows():
            # Omitir mercados sin objetivos para no romper la interfaz
            if row['Objetivos'] == 0:
                continue
                
            col_nom, col_val = st.columns(2)
            col_nom.markdown(f"**{row['MERCADO']}**")
            col_val.markdown(f"<p style='text-align: right; margin:0;'>{row['Producción']:,.0f} / {row['Objetivos']:,.0f}</p>", unsafe_allow_html=True)
            
            # 1. ASIGNAR COLOR DINÁMICO SEGÚN EL PORCENTAJE
            if row['Porcentaje_Ind'] >= 100:
                color_barra = "#4cd964"  # Verde para éxito
                texto_estado = "Sobre objetivo"
            elif row['Porcentaje_Ind'] >= 80:
                color_barra = "#ffcc00"  # Amarillo/Naranja claro para advertencia
                texto_estado = "Bajo objetivo"
            else:
                color_barra = "#ff3b30"  # Rojo para estado crítico
                texto_estado = "Bajo objetivo"
                
            # Limitar el porcentaje visual entre 0% y 100% para que la barra HTML no se desborde
            ancho_barra = min(max(row['Porcentaje_Ind'], 0), 100)
            
            # 2. REMPLAZAR ST.PROGRESS POR BARRA HTML PERSONALIZADA
            html_barra = f"""
            <div style="background-color: #262730; width: 100%; border-radius: 4px; height: 8px; margin: 8px 0;">
                <div style="background-color: {color_barra}; width: {ancho_barra}%; border-radius: 4px; height: 100%;"></div>
            </div>
            """
            # 3. MOSTRAR EL TEXTO INFERIOR CON EL MISMO COLOR
            st.markdown(f"<p style='color: {color_barra}; font-size: 14px; margin-top: -5px; margin-bottom: 20px;'>{row['Porcentaje_Ind']}% — {texto_estado}</p>", unsafe_allow_html=True)
            st.markdown(html_barra, unsafe_allow_html=True)

    # --- 5. DISEÑO DE INTERFAZ: TABLA RESUMEN INFERIOR ---

    with tarjeta_completa:
        # 1. Crear el DataFrame base para la tabla
        df_tabla = df[["MERCADO", "Objetivos", "Producción"]].copy()
        df_tabla["Diferencia"] = df_tabla["Producción"] - df_tabla["Objetivos"]

        # 2. Calcular e insertar la fila de totales
        fila_total = pd.DataFrame([{
            "MERCADO": "Total",
            "Objetivos": total_obj,
            "Producción": total_real,
            "Diferencia": total_dif
        }])
        df_final = pd.concat([df_tabla, fila_total], ignore_index=True)

        # 3. FUNCIONES DE ESTILO CSS
        def colorear_diferencia(valor):
            if valor > 0:
                return "color: #4cd964; font-weight: bold;"  # Verde si es positivo
            elif valor < 0:
                return "color: #ff3b30; font-weight: bold;"  # Rojo si es negativo
            else:
                return "color: #ffffff;"  # Blanco si es exactamente cero

        # NUEVA FUNCIÓN: Pone en negrita toda la fila si corresponde al Total
        def destacar_total(fila):
            if fila["MERCADO"] == "Total":
                return ["font-weight: bold; border-top: 1px solid #444;"] * len(fila)
            return [""] * len(fila)

        # 4. APLICAR CONFIGURACIÓN COMPLETA AL STYLER (Una sola cadena de métodos)
        df_estilizado = (
            df_final.style
            # Aplicar colores a la columna Diferencia
            .map(colorear_diferencia, subset=["Diferencia"])
            # Aplicar la negrita a la fila completa del Total
            .apply(destacar_total, axis=1)
            # Diseñar los encabezados de las columnas (th)
            .set_table_styles([
                {
                    'selector': 'th',
                    'props': [
                        ('background-color', '#1e2029'),
                        ('color', '#ffffff'),
                        ('font-weight', 'bold'),
                        ('text-align', 'center'),
                        ('font-size', '15px')
                    ]
                }
            ])
            # Formatear visualmente los números
            .format({
                "Objetivos": "{:,.0f}",
                "Producción": "{:,.0f}",
                "Diferencia": lambda x: f"+{x:,.0f}" if x > 0 else f"{x:,.0f}"
            })
        )

        # 5. DESPLEGAR LA TABLA ESTILIZADA EN STREAMLIT
        st.dataframe(df_estilizado, hide_index=True, use_container_width=True)
